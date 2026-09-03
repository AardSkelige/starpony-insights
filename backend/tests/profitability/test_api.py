"""Прибыльность: доступ к разделу, разбор запроса, контракт ответа.

Контракт проверяется через API, а не через сервис. Причина — дефект соседней
страницы: доля больше единицы не влезала в `DecimalField(9, 8)`, и весь ответ
уходил пятисотой. Тест на сервисе такого не видит вовсе.

**Себестоимость и маржа видны только тем, кому открыта эта страница.**
Решение владельца 30.08, полутонов нет: страница открыта — числа видны,
закрыта — не видны вовсе.
"""

from io import BytesIO
from urllib.parse import unquote

import pytest

from api.profitability.serializers import ProfitabilityQuerySerializer

from .conftest import moscow, position

pytestmark = pytest.mark.django_db

PAGE_KEY = "profitability"
URL = "/api/profitability/"


@pytest.fixture
def sales(product, make_product, make_profit_day, make_demand, make_contract):
    """Выборка, в которой есть всё, о чём страница обязана уметь говорить."""
    make_profit_day(quantity="10", revenue_kopecks=100_000, cost_kopecks=30_000,
                    marketplace_quantity="4", marketplace_revenue_kopecks=60_000,
                    marketplace_cost_kopecks=10_000)
    gift = make_demand(moment=moscow(2026, 7, 15))
    position(gift, product, 2, 0)

    consigned = make_product(name="Воск", article="400.003.15",
                             folder="Готовая продукция/Амуниция для лошадей")
    contract = make_contract()
    shipment = make_demand(moment=moscow(2026, 7, 15), contract=contract)
    position(shipment, consigned, 8, 20_000)
    make_profit_day(product=consigned, quantity="3", revenue_kopecks=6_000,
                    cost_kopecks=9_000)
    return product


class TestAccess:
    def test_requires_login(self, client):
        assert client.get(URL).status_code == 401

    def test_requires_this_page(self, client, make_user):
        """Права соседнего раздела не открывают этот.

        Здесь это не формальность: страница показывает себестоимость,
        и открыть её случайно значит показать её всем вошедшим.
        """
        client.force_login(make_user(pages=["channels"]))
        assert client.get(URL).status_code == 403

    def test_allows_granted_user(self, client, make_user):
        client.force_login(make_user(pages=[PAGE_KEY]))
        assert client.get(URL).status_code == 200


class TestQuery:
    def test_basis_is_part_of_the_contract(self):
        """Переключатель базы живёт в адресной строке — ссылку можно переслать."""
        fields = ProfitabilityQuerySerializer().fields

        assert "basis" in fields
        assert "with_free" in fields

    def test_unknown_basis_is_rejected(self, client, make_user):
        """Неизвестная база — четырёхсотая, а не молчаливое умолчание.

        Иначе ссылка с опечаткой показала бы другие числа, ничем
        не отличаясь на вид.
        """
        client.force_login(make_user(pages=[PAGE_KEY]))

        assert client.get(URL, {"basis": "invented"}).status_code == 400

    def test_period_backwards_is_rejected(self, client, make_user):
        client.force_login(make_user(pages=[PAGE_KEY]))

        response = client.get(
            URL, {"date_from": "2026-08-01", "date_to": "2026-07-01"}
        )

        assert response.status_code == 400


class TestContract:
    @pytest.fixture
    def payload(self, client, make_user, sales):
        client.force_login(make_user(pages=[PAGE_KEY]))
        response = client.get(URL, {"page_size": 200})
        assert response.status_code == 200
        return response.json()

    def test_row_carries_the_parts_of_its_margin(self, payload):
        """Расчётное число приходит составляющими, а не готовым текстом.

        Формула на экране собирается из полученного (`CLAUDE.md` §4);
        рубли и проценты существуют только на слое отображения.
        """
        row = payload["results"][0]

        for field in ("quantity", "revenue_kopecks", "cost_kopecks",
                      "profit_kopecks", "margin", "unit_cost_kopecks",
                      "cost_is_estimated"):
            assert field in row, field

    def test_row_names_what_is_left_out(self, payload):
        """Подарки и товар на реализации — в самой строке, а не только в сводке.

        Иначе строка «продано 8 шт» рядом с «отгружено 10» выглядит ошибкой.
        """
        row = payload["results"][0]

        for field in ("free_quantity", "free_cost_kopecks", "unsold_quantity",
                      "unsold_kopecks", "shipped_quantity", "sold_quantity"):
            assert field in row, field

    def test_marketplace_margin_comes_next_to_the_direct_one(self, payload):
        """Два числа рядом: завышенное и настоящее."""
        marketplaces = payload["marketplaces"]

        assert marketplaces["marketplace_margin"] is not None
        assert marketplaces["direct_margin"] is not None

    def test_coverage_says_what_is_outside_the_margin(self, payload):
        coverage = payload["coverage"]

        assert coverage["basis"] == "sold"
        assert coverage["with_free"] is False
        assert coverage["sold_revenue_kopecks"] != coverage["shipped_revenue_kopecks"]

    def test_huge_share_does_not_break_the_response(
        self, client, make_user, make_product, make_profit_day
    ):
        """Доля в разы больше единицы обязана влезать в поле.

        На соседней странице такая доля не влезла в `DecimalField(9, 8)` —
        туда помещается лишь 9,99999999, — и весь ответ уходил пятисотой.

        Возникает она штатно, без экзотики: убыточный товар почти съедает
        прибыль остальных, знаменатель становится маленьким, и доля лидера
        уходит за тысячу процентов. Доля 1,4 такой проверкой не была бы —
        она влезает и в узкое поле, на чём этот тест и был пойман
        проверкой на покраснение.
        """
        winner = make_product(name="Репеллент", article="300.001.05")
        make_profit_day(product=winner, revenue_kopecks=200_000, cost_kopecks=100_000)
        loser = make_product(name="Воск", article="400.003.15")
        make_profit_day(product=loser, revenue_kopecks=10_000, cost_kopecks=105_000)

        client.force_login(make_user(pages=[PAGE_KEY]))
        response = client.get(URL, {"page_size": 200})

        assert response.status_code == 200
        shares = [row["profit_share"] for row in response.json()["results"]]
        # 100 000 прибыли лидера против 5 000 по всей выборке — двадцать раз.
        assert max(float(share) for share in shares if share is not None) == 20


class TestExport:
    """Выгрузка отдаёт отобранное целиком, а не видимую страницу."""

    def test_returns_a_workbook(self, client, make_user, sales):
        client.force_login(make_user(pages=[PAGE_KEY]))

        response = client.get(f"{URL}xlsx/")

        assert response.status_code == 200
        assert response["Content-Type"].startswith(
            "application/vnd.openxmlformats"
        )

    def test_carries_the_whole_selection_not_one_page(
        self, client, make_user, make_product, make_profit_day
    ):
        """Строк больше, чем влезает на страницу, — в файле все.

        Высота страницы подрезается потолком в 200 (`page_bounds`), и позови
        выгрузка `page`, хвост выборки исчез бы из файла молча.
        """
        from openpyxl import load_workbook

        for i in range(1, 13):
            product = make_product(name=f"Товар {i}", article=f"300.{i:03d}.05")
            make_profit_day(product=product, revenue_kopecks=10_000 * i,
                            cost_kopecks=1_000 * i)

        client.force_login(make_user(pages=[PAGE_KEY]))
        # Высота страницы по умолчанию — десять строк.
        response = client.get(f"{URL}xlsx/")

        book = load_workbook(BytesIO(b"".join(response.streaming_content)))
        sheet = book.worksheets[0]
        # 12 товаров + шапка + строка итога.
        assert sheet.max_row == 14

    def test_basis_reaches_the_file_name_and_the_sheet(
        self, client, make_user, sales
    ):
        """Два файла за один период обязаны различаться на вид.

        Иначе они расходятся на 281 126 ₽, а объяснения в них нет ни строчки.
        """
        from openpyxl import load_workbook

        client.force_login(make_user(pages=[PAGE_KEY]))
        response = client.get(f"{URL}xlsx/", {"basis": "shipped"})

        # Имя едет процентным кодированием — сравнивать надо расшифрованное.
        disposition = unquote(response["Content-Disposition"])
        assert "отгружено" in disposition.lower()
        book = load_workbook(BytesIO(b"".join(response.streaming_content)))
        assert book.worksheets[0].title == "Отгружено"
