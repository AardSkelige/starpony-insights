"""Выгрузка «Товаров в отгрузках» в XLSX."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from api.shipments.services import excel, products
from tests.shipments.conftest import moscow, position

pytestmark = pytest.mark.django_db

URL = "/api/shipments/products/xlsx/"
PAGE_KEY = "shipments-products"


def sheet_of(filters: products.Filters):
    return load_workbook(excel.build(filters)).active


def column_of(key: str) -> int:
    """Номер колонки по её ключу — чтобы тест не считал буквы руками."""
    return next(
        index for index, (_, __, name) in enumerate(excel.COLUMNS, start=1) if name == key
    )


def test_export_covers_every_page_not_just_the_first(make_product, make_demand):
    """В файл уходит вся выборка, а не видимая страница.

    Человек отбирает период и жмёт «Экспорт», ожидая получить отобранное.
    Пятьдесят строк вместо трёхсот — потеря, которую он заметит не сразу.
    """
    count = excel.CHUNK + 5
    for index in range(count):
        position(make_demand(), make_product(code=f"2-{index:04d}"), "1.000", 10000)

    sheet = sheet_of(products.Filters(page_size=10))

    # Строки: шапка + товары + итог.
    assert sheet.max_row == count + 2


def test_export_respects_filters(make_product, make_demand, channel):
    """Выгружается то, что отобрано, а не всё подряд."""
    product = make_product()
    position(make_demand(channel=channel), product, "1.000", 10000)
    position(make_demand(), make_product(code="2-999"), "1.000", 70000)

    sheet = sheet_of(products.Filters(channel_id=channel.id))

    assert sheet.max_row == 3
    assert sheet.cell(row=2, column=column_of("code")).value == "2-001"


def test_export_writes_rubles_not_kopecks(make_product, make_demand):
    """В файле рубли: копейки в ячейке читаются как сумма в сто раз больше."""
    position(make_demand(), make_product(), "2.000", 23153038)

    sheet = sheet_of(products.Filters())

    assert sheet.cell(row=2, column=column_of("revenue")).value == pytest.approx(231530.38)


def test_export_keeps_free_quantity_separate(make_product, make_demand):
    """Бесплатные штуки видны отдельной колонкой, как и на экране."""
    product = make_product()
    position(make_demand(), product, "4.000", 40000)
    position(make_demand(), product, "1.000", 0)

    sheet = sheet_of(products.Filters())

    assert sheet.cell(row=2, column=column_of("quantity")).value == pytest.approx(5)
    assert sheet.cell(row=2, column=column_of("free_quantity")).value == pytest.approx(1)
    # Цена без бесплатных — 100 ₽, со всеми — 80 ₽.
    assert sheet.cell(row=2, column=column_of("avg_price")).value == pytest.approx(80)
    assert sheet.cell(row=2, column=column_of("avg_price_paid")).value == pytest.approx(100)


def test_export_ends_with_totals(make_product, make_demand):
    """Последняя строка — итог: без него файл приходится складывать вручную."""
    position(make_demand(), make_product(code="2-001"), "2.000", 30000)
    position(make_demand(), make_product(code="2-002"), "3.000", 20000)

    sheet = sheet_of(products.Filters())
    last = sheet.max_row

    assert "Итого" in str(sheet.cell(row=last, column=column_of("name")).value)
    assert sheet.cell(row=last, column=column_of("quantity")).value == pytest.approx(5)
    assert sheet.cell(row=last, column=column_of("revenue")).value == pytest.approx(500)


def test_export_headers_match_the_screen(make_product, make_demand):
    """Подписи и порядок колонок совпадают с экраном.

    Файл, где колонки идут иначе, заставляет сверять их глазами — и именно
    там появляются ошибки переноса чисел в отчёт.
    """
    position(make_demand(), make_product(), "1.000", 10000)

    sheet = sheet_of(products.Filters())
    headers = [sheet.cell(row=1, column=i).value for i in range(1, len(excel.COLUMNS) + 1)]

    assert headers == [title for title, _, __ in excel.COLUMNS]
    assert "в т.ч. даром" in headers
    assert "Средняя цена продажи, ₽" in headers


def test_export_freezes_the_header(make_product, make_demand):
    """Шапка закреплена: иначе на сороковой строке не видно, что за колонка."""
    position(make_demand(), make_product(), "1.000", 10000)

    assert sheet_of(products.Filters()).freeze_panes == "A2"


def test_export_requires_the_page(client, make_user):
    """Выгрузка закрыта той же страницей, что и сама таблица.

    Иначе экспорт становится обходом прав: данные те же, а проверка другая.
    """
    client.force_login(make_user(pages=["deadlines"]))

    assert client.get(URL).status_code == 403


def test_export_returns_a_spreadsheet(client, make_user, make_product, make_demand):
    position(make_demand(moment=moscow(2026, 6, 15)), make_product(), "1.000", 10000)
    client.force_login(make_user(pages=[PAGE_KEY]))

    response = client.get(URL)

    assert response.status_code == 200
    assert "spreadsheetml" in response["Content-Type"]
    assert "attachment" in response["Content-Disposition"]
    # Файл действительно открывается, а не просто отдаётся с нужным типом.
    book = load_workbook(BytesIO(b"".join(response.streaming_content)))
    assert book.active.title == "Товары в отгрузках"


def test_export_rejects_reversed_period(client, make_user):
    client.force_login(make_user(pages=[PAGE_KEY]))

    response = client.get(URL, {"date_from": "2026-07-01", "date_to": "2026-06-01"})

    assert response.status_code == 400


def test_file_name_carries_the_period(client, make_user, make_product, make_demand):
    """В имени файла — период данных, а не только дата выгрузки.

    Две выборки за разные периоды, скачанные в один день, получили бы
    одинаковое имя, и в папке «Загрузки» их не различить.
    """
    position(make_demand(moment=moscow(2026, 6, 15)), make_product(), "1.000", 10000)
    client.force_login(make_user(pages=[PAGE_KEY]))

    response = client.get(URL, {"date_from": "2026-06-01", "date_to": "2026-06-30"})

    disposition = response["Content-Disposition"]
    assert "01.06.2026" in disposition
    assert "30.06.2026" in disposition


def test_file_name_says_when_no_period_is_chosen(client, make_user, make_product, make_demand):
    """Без выбранного периода имя говорит об этом прямо, а не молчит."""
    position(make_demand(), make_product(), "1.000", 10000)
    client.force_login(make_user(pages=[PAGE_KEY]))

    assert "%D0%B2%D0%B5%D1%81%D1%8C" in client.get(URL)["Content-Disposition"]


def test_totals_row_declines_the_noun(make_product, make_demand):
    """«1 наименование», а не «1 наименований».

    Число в итоге приходит из выборки и бывает любым, поэтому слово рядом
    с ним обязано склоняться. Фраза собирается один раз — и ошибка в ней
    видна человеку в каждой выгрузке.
    """
    position(make_demand(), make_product(), "1.000", 10000)

    sheet = sheet_of(products.Filters())
    last = sheet.max_row

    assert sheet.cell(row=last, column=column_of("name")).value == "Итого · 1 наименование"


def test_export_names_the_consignment_part(make_product, make_demand, commission):
    """Выручка в файле обязана нести ту же оговорку, что полоса на экране.

    Полосы в книге нет, а колонку «Выручка» складывают сводной таблицей
    первым делом. Без соседнего числа файл утверждает, что товар,
    лежащий у комиссионера, уже продан: на боевых так сложились бы
    281 126 ₽ (`CLAUDE.md` §8.0).
    """
    product = make_product()
    position(make_demand(contract=commission), product, "1.000", 30000)
    position(make_demand(), product, "1.000", 20000)

    sheet = sheet_of(products.Filters())
    last = sheet.max_row

    assert sheet.cell(row=2, column=column_of("revenue")).value == pytest.approx(500)
    assert sheet.cell(row=2, column=column_of("consignment")).value == pytest.approx(300)
    assert sheet.cell(row=last, column=column_of("consignment")).value == pytest.approx(300)


def test_export_writes_zero_when_nothing_is_consigned(make_product, make_demand):
    """Ноль, а не пустая ячейка: пустая читается как «не посчитали»."""
    position(make_demand(), make_product(), "1.000", 20000)

    sheet = sheet_of(products.Filters())

    assert sheet.cell(row=2, column=column_of("consignment")).value == pytest.approx(0)
