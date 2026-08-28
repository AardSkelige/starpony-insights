"""Выгрузка «Материалов в отгрузках» в XLSX."""

import pytest
from openpyxl import load_workbook

from api.shipments.services import excel_materials, materials
from core.models import ProductKind
from tests.shipments.conftest import moscow, position

pytestmark = pytest.mark.django_db

URL = "/api/shipments/materials/xlsx/"
PAGE_KEY = "shipments-materials"


@pytest.fixture
def sold(make_product, make_plan, make_demand, make_supply):
    bottled = make_product("Шампунь 500 мл", article="100.001", code="2-001")
    water = make_product("Вода дистиллированная", article="W-1", code="9-001")
    make_plan("Розлив", bottled, output=1, materials=[(water, 50)])
    make_supply(water, "2.50", moment=moscow(2026, 6, 15))

    delivery = make_product("Доставка", article="", code="")
    delivery.kind = ProductKind.SERVICE
    delivery.save(update_fields=["kind"])

    document = make_demand()
    position(document, bottled, "10", 500_00)
    position(document, delivery, "1", 300_00)
    return {"bottled": bottled, "water": water, "delivery": delivery}


def book_of(filters: materials.Filters):
    return load_workbook(excel_materials.build(filters))


def column_of(key: str, columns) -> int:
    """Номер колонки по её ключу — чтобы тест не считал буквы руками."""
    return next(
        index for index, (_, __, name) in enumerate(columns, start=1) if name == key
    )


def test_two_sheets_keep_services_out_of_the_sum(sold):
    """Доставка живёт на своём листе, а не в сумме сырья.

    В одном списке её кто-нибудь обязательно сложит вместе с материалами,
    и «сырья на 399 686 ₽» перестанет значить то, что написано.
    """
    book = book_of(materials.Filters())
    assert book.sheetnames == ["Материалы в отгрузках", "Без техкарты"]

    names = column_of("name", excel_materials.COLUMNS)
    first = book["Материалы в отгрузках"]
    assert [first.cell(row=line, column=names).value for line in (1, 2)] == [
        "Материал",
        "Вода дистиллированная",
    ]

    second = book["Без техкарты"]
    kind = column_of("kind", excel_materials.WITHOUT_PLAN_COLUMNS)
    assert second.cell(row=2, column=names).value == "Доставка"
    assert second.cell(row=2, column=kind).value == "Услуга"


def test_second_sheet_exists_even_when_empty(make_product, make_plan, make_demand):
    """Лист есть всегда: его отсутствие человек прочтёт как «всё развернулось».

    А это разные вещи, и проверить их в файле нечем.
    """
    bottled = make_product("Шампунь 500 мл", article="100.001", code="2-001")
    water = make_product("Вода", article="W-1", code="9-001")
    make_plan("Розлив", bottled, output=1, materials=[(water, 50)])
    position(make_demand(), bottled, "1", 50_00)

    book = book_of(materials.Filters())
    assert "Без техкарты" in book.sheetnames
    assert book["Без техкарты"].max_row == 1, "остались лишние строки"


def test_export_covers_the_whole_selection(sold, make_product, make_plan, make_demand):
    """В файл уходит вся выборка, а не видимая страница."""
    extra = make_product("Кондиционер 500 мл", article="100.002", code="2-002")
    others = [
        make_product(f"Сырьё {index}", article=f"S-{index}", code=f"9-1{index:02d}")
        for index in range(12)
    ]
    make_plan("Розлив-2", extra, output=1, materials=[(item, 1) for item in others])
    position(make_demand(), extra, "1", 100_00)

    sheet = book_of(materials.Filters(page_size=2))["Материалы в отгрузках"]
    # Шапка + тринадцать материалов + итог.
    assert sheet.max_row == 15


def test_export_respects_filters(sold, make_demand, make_channel):
    """Выгружается то, что отобрано."""
    ozon = make_channel("Озон")
    position(make_demand(channel=ozon), sold["bottled"], "4", 200_00)

    sheet = book_of(materials.Filters(channel_id=ozon.pk))["Материалы в отгрузках"]
    quantity = column_of("quantity", excel_materials.COLUMNS)
    assert sheet.cell(row=2, column=quantity).value == 200.0


def test_price_and_cost_reach_the_file(sold):
    """Цена и её дата уходят в файл: без них стоимость нечем проверить."""
    sheet = book_of(materials.Filters())["Материалы в отгрузках"]
    price = column_of("price", excel_materials.COLUMNS)
    when = column_of("price_date", excel_materials.COLUMNS)
    cost = column_of("cost", excel_materials.COLUMNS)

    assert sheet.cell(row=2, column=price).value == pytest.approx(0.025)
    assert sheet.cell(row=2, column=when).value == "15.06.2026"
    assert sheet.cell(row=2, column=cost).value == pytest.approx(12.5)


def test_totals_row_closes_the_table(sold):
    sheet = book_of(materials.Filters())["Материалы в отгрузках"]
    name = column_of("name", excel_materials.COLUMNS)
    assert "Итого" in sheet.cell(row=sheet.max_row, column=name).value


def test_file_name_says_the_period(client, make_user, sold):
    """Имя файла говорит, что внутри и когда снято."""
    client.force_login(make_user(pages=[PAGE_KEY]))
    response = client.get(URL, {"date_from": "2026-06-01", "date_to": "2026-06-30"})

    assert response.status_code == 200
    disposition = response.headers["Content-Disposition"]
    assert "01.06.2026" in disposition and "30.06.2026" in disposition
