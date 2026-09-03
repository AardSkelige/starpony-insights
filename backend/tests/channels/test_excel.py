"""Выгрузка «Каналов продаж» в XLSX.

Проверяется здесь одно — что число в файле означает то же, что на экране.
Оговорка про товар на реализации живёт на странице полосой над таблицей,
а в книге полосы нет: остаётся колонка, и без неё файл утверждает, что
товар, лежащий у комиссионера, уже продан. На боевых так сложились бы
281 126 ₽ — 87 % выручки крупнейшего канала (`CLAUDE.md` §8.0).
"""

import pytest
from openpyxl import load_workbook

from api.channels.services import channels as service, excel
from tests.channels.conftest import position

pytestmark = pytest.mark.django_db


def sheet_of(filters: service.Filters):
    return load_workbook(excel.build(filters)).worksheets[0]


def column_of(key: str) -> int:
    return next(
        index
        for index, (_, __, name) in enumerate(excel.COLUMNS, start=1)
        if name == key
    )


def test_headers_match_the_screen(make_demand, make_product):
    """Порядок и подписи те же, что в таблице: иначе колонки сверяют глазами."""
    position(make_demand(), make_product(), "1.000", 10000)

    sheet = sheet_of(service.Filters())
    headers = [
        sheet.cell(row=1, column=i).value for i in range(1, len(excel.COLUMNS) + 1)
    ]

    assert headers == [title for title, _, __ in excel.COLUMNS]


def test_consignment_stands_next_to_revenue(
    make_demand, make_product, make_contract
):
    """Сколько из выручки канала — товар, который ещё не продан.

    Рядом с выручкой, а не в конце: через шесть колонок число к выручке
    уже не относят.
    """
    product = make_product()
    position(
        make_demand(total_kopecks=300_00, contract=make_contract()), product, "1", 300_00
    )
    position(make_demand(total_kopecks=200_00), product, "1", 200_00)

    sheet = sheet_of(service.Filters())
    last = sheet.max_row

    assert sheet.cell(row=2, column=column_of("revenue")).value == pytest.approx(500)
    assert sheet.cell(row=2, column=column_of("consignment")).value == pytest.approx(300)
    assert sheet.cell(row=last, column=column_of("consignment")).value == pytest.approx(
        300
    )


def test_sales_contract_is_not_consignment(
    make_demand, make_product, make_contract
):
    """Договор купли-продажи — обычная продажа, ноль в колонке."""
    from core.models import ContractType

    position(
        make_demand(
            total_kopecks=200_00,
            contract=make_contract(contract_type=ContractType.SALES),
        ),
        make_product(),
        "1",
        200_00,
    )

    sheet = sheet_of(service.Filters())

    assert sheet.cell(row=2, column=column_of("consignment")).value == pytest.approx(0)
