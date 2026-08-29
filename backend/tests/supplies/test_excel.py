"""Выгрузка «Материалы в приёмках»: два листа, и они про одно и то же.

Файл открывают, чтобы складывать. Расхождение между листами или между
подвалом и колонкой заметят раньше нас.
"""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from api.supplies.services import excel, materials
from tests.supplies.conftest import moscow, position

pytestmark = pytest.mark.django_db


def book(filters=None):
    return load_workbook(excel.build(filters or materials.Filters()))


def values(sheet, key: str, columns) -> list:
    """Колонка листа по ключу из описания — без счёта позиций глазами."""
    index = next(
        number for number, (_, __, name) in enumerate(columns, start=1) if name == key
    )
    return [row[index - 1].value for row in sheet.iter_rows(min_row=2)]


@pytest.fixture
def two_materials(make_supply, make_product):
    scent = make_product("Отдушка Хлопок", article="1.001", code="1-001")
    bottle = make_product("Флакон 500 мл", article="2.001", code="2-001")
    position(make_supply(moment=moscow(2026, 4, 19)), scent, 10, "754")
    position(make_supply(moment=moscow(2026, 6, 9)), bottle, 100, "2505")
    position(make_supply(moment=moscow(2026, 8, 18)), bottle, 100, "3105")
    return {"scent": scent, "bottle": bottle}


class TestSheets:
    def test_both_sheets_exist(self, two_materials):
        """Лист закупок создаётся всегда, даже пустым.

        Его отсутствие человек прочтёт как «закупок не было», а это
        другое утверждение — и проверить его будет нечем.
        """
        assert book().sheetnames == ["Материалы в приёмках", "Закупки"]

    def test_purchase_sheet_has_a_row_per_supply(self, two_materials):
        sheet = book()["Закупки"]
        assert sheet.max_row == 4  # шапка и три приёмки

    def test_purchase_sheet_is_chronological(self, two_materials):
        """Лист читают как ленту закупок, а не как список товаров."""
        sheet = book()["Закупки"]
        assert values(sheet, "date", excel.PURCHASE_COLUMNS) == [
            "19.04.2026",
            "09.06.2026",
            "18.08.2026",
        ]


class TestWholeSelection:
    def test_export_is_not_cut_to_a_page(self, make_supply, make_product):
        """Файл берёт всю выборку, а не первую страницу.

        Тот же дефект уже был на соседней странице: выгрузка обрезалась
        на 200 строках, а строка итога считалась по всем — и заметно это
        стало бы только после расширения периода.
        """
        for index in range(15):
            position(
                make_supply(),
                make_product(f"Материал {index}", code=f"{index}"),
                1,
                "100",
            )

        sheet = book(materials.Filters(page_size=5))["Материалы в приёмках"]
        assert sheet.max_row == 17  # шапка, пятнадцать строк и итог

    def test_totals_row_matches_the_shown_column(self, two_materials):
        """Подвал считается по строкам файла, а не по всей выборке."""
        sheet = book(materials.Filters(search="отдушка"))["Материалы в приёмках"]
        amounts = values(sheet, "amount", excel.COLUMNS)
        assert amounts[-1] == sum(value for value in amounts[:-1] if value)

    def test_search_narrows_both_sheets(self, two_materials):
        """Два листа одного файла обязаны быть про одно и то же.

        Оставь мы на втором все закупки, сумма по нему не сошлась бы
        с суммой по первому — и какой из двух верен, по файлу не понять.
        """
        whole = book(materials.Filters(search="отдушка"))
        assert whole["Материалы в приёмках"].max_row == 3  # шапка, строка, итог
        assert whole["Закупки"].max_row == 2  # шапка и одна приёмка


class TestCells:
    def test_free_quantity_is_blank_when_nothing_was_free(self, two_materials):
        """Пусто, а не ноль: даром не приходило ничего — это не «пришло ноль»."""
        sheet = book()["Материалы в приёмках"]
        assert values(sheet, "free_quantity", excel.COLUMNS)[:2] == [None, None]

    def test_price_columns_carry_numbers_not_text(self, two_materials):
        """Цена уходит числом: файл открывают ради сводных таблиц."""
        sheet = book()["Материалы в приёмках"]
        prices = values(sheet, "last_price", excel.COLUMNS)
        assert prices[0] == pytest.approx(31.05)

    def test_unpriced_material_has_no_price(self, make_supply, make_product):
        label = make_product("Этикетка")
        position(make_supply(), label, 100, "0")

        sheet = book()["Материалы в приёмках"]
        assert values(sheet, "avg_price", excel.COLUMNS)[0] is None
        assert values(sheet, "change", excel.COLUMNS)[0] is None

    def test_response_streams_a_readable_book(self, client, make_user, two_materials):
        client.force_login(make_user(pages=["supplies-materials"]))
        response = client.get("/api/supplies/materials/xlsx/")
        whole = load_workbook(BytesIO(b"".join(response.streaming_content)))
        assert whole.sheetnames == ["Материалы в приёмках", "Закупки"]


class TestFooterIsAboutShownRows:
    """Все числа подвала — про одно множество, и это показанные строки.

    Сумма бралась из итога, а приёмки с поставщиками — из охвата: подвал
    при выгрузке с поиском читался как «Итого · 1 материал … Закупок 3 ·
    Поставщиков 2», где 3 и 2 описывают всю выборку. Дробь, где числитель
    от найденного, а знаменатель от всего, выглядит обычной и врёт.
    """

    def test_supply_count_follows_the_search(self, two_materials):
        sheet = book(materials.Filters(search="отдушка"))["Материалы в приёмках"]
        footer = values(sheet, "supplies_count", excel.COLUMNS)[-1]

        # Отдушка пришла одной приёмкой из трёх в выборке.
        assert footer == 1

    def test_supplier_count_follows_the_search(
        self, make_supply, make_product, make_supplier
    ):
        scent = make_product("Отдушка Хлопок", article="1.001", code="1-001")
        bottle = make_product("Флакон 500 мл", article="2.001", code="2-001")
        position(make_supply(agent=make_supplier("Первый")), scent, 10, "754")
        position(make_supply(agent=make_supplier("Второй")), bottle, 10, "2505")

        sheet = book(materials.Filters(search="отдушка"))["Материалы в приёмках"]
        assert values(sheet, "suppliers_count", excel.COLUMNS)[-1] == 1

    def test_whole_selection_counts_everything(self, two_materials):
        """Без поиска подвал описывает всю выборку — как и раньше."""
        sheet = book()["Материалы в приёмках"]

        assert values(sheet, "supplies_count", excel.COLUMNS)[-1] == 3
        assert values(sheet, "suppliers_count", excel.COLUMNS)[-1] == 1

    def test_one_supply_of_several_materials_counts_once(
        self, make_supply, make_product
    ):
        """Приёмка с тремя материалами — одна закупка, а не три.

        Сложить `supplies_count` по строкам нельзя: одна приёмка приносит
        несколько наименований и была бы посчитана столько раз, сколько
        в ней позиций.
        """
        supply = make_supply()
        for index in range(3):
            position(supply, make_product(f"Материал {index}", code=f"{index}"), 10, "100")

        sheet = book()["Материалы в приёмках"]
        assert values(sheet, "supplies_count", excel.COLUMNS)[-1] == 1
