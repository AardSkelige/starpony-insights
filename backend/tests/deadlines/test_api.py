"""Сроки оплаты: доступ к разделу, разбор запроса, контракт ответа.

Контракт проверяется через API, а не через сервис. Причина — дефект соседней
страницы: доля больше единицы не влезала в `DecimalField(9, 8)`, и весь ответ
уходил пятисотой. Тест на сервисе такого не видит вовсе.
"""

import pytest

from api.deadlines.serializers import DeadlinesQuerySerializer
from core.models import DocumentKind

pytestmark = pytest.mark.django_db

PAGE_KEY = "deadlines"
URL = "/api/deadlines/"


@pytest.fixture
def owing(make_agent, make_document, make_channel, make_contract):
    buyer = make_agent("ООО «ПМТ»")
    ozon = make_agent("ООО «Интернет Решения»", tags=["маркетплейсы"])
    caprioli = make_agent("КРМОО «Каприоль»")
    contract = make_contract(caprioli)

    make_document(agent=buyer, age_days=5, total_kopecks=100_000,
                  sales_channel=make_channel("ХорсСмарт"))
    make_document(agent=buyer, age_days=70, total_kopecks=200_000)
    make_document(agent=ozon, age_days=20, total_kopecks=900_000)
    make_document(agent=caprioli, contract=contract, total_kopecks=400_000)
    make_document(agent=caprioli, contract=contract,
                  kind=DocumentKind.COMMISSION_REPORT, total_kopecks=80_000)
    return buyer


class TestAccess:
    def test_requires_login(self, client):
        assert client.get(URL).status_code == 401

    def test_requires_this_page(self, client, make_user):
        """Права соседнего раздела не открывают этот."""
        client.force_login(make_user(pages=["channels"]))
        assert client.get(URL).status_code == 403

    def test_allows_granted_user(self, client, make_user):
        client.force_login(make_user(pages=[PAGE_KEY]))
        assert client.get(URL).status_code == 200


class TestQuery:
    def test_period_is_not_part_of_the_contract(self):
        """Периода у страницы нет, и в схеме его быть не должно.

        Долг — состояние на сегодня, а не итог за отрезок. Оставь мы поля
        «на всякий случай», ссылка с `?date_from=…` выглядела бы рабочей
        и показывала не то, что обещает.
        """
        fields = DeadlinesQuerySerializer().fields

        assert "date_from" not in fields
        assert "date_to" not in fields
        assert "search" in fields
        assert "ordering" in fields

    def test_unknown_ordering_is_rejected(self, client, make_user):
        client.force_login(make_user(pages=[PAGE_KEY]))
        assert client.get(URL, {"ordering": "-revenue"}).status_code == 400

    def test_page_size_beyond_the_ceiling_is_rejected(self, client, make_user):
        client.force_login(make_user(pages=[PAGE_KEY]))
        assert client.get(URL, {"page_size": 100_000}).status_code == 400


class TestPayload:
    def test_three_sums_arrive_apart(self, client, make_user, owing):
        """Дебиторка, площадки и реализация — три поля, а не одно число.

        Сложи их — получится 1 680 000 «долга» вместо 380 000, которые
        действительно должны.
        """
        client.force_login(make_user(pages=[PAGE_KEY]))

        body = client.get(URL).json()

        assert body["totals"]["debt_kopecks"] == 380_000
        assert body["coverage"]["marketplace_kopecks"] == 900_000
        assert body["coverage"]["consignment_kopecks"] == 400_000

    def test_marketplaces_come_beside_the_rows(self, client, make_user, owing):
        client.force_login(make_user(pages=[PAGE_KEY]))

        body = client.get(URL).json()

        assert [row["name"] for row in body["marketplaces"]] == [
            "ООО «Интернет Решения»"
        ]
        assert "ООО «Интернет Решения»" not in [row["name"] for row in body["results"]]

    def test_aging_arrives_with_all_four_shelves(self, client, make_user, owing):
        client.force_login(make_user(pages=[PAGE_KEY]))

        body = client.get(URL).json()

        assert [shelf["key"] for shelf in body["aging"]] == [
            "fresh", "recent", "stale", "old",
        ]

    def test_deferral_is_null_not_zero(self, client, make_user, owing):
        """Ноль означал бы «платят в день отгрузки» — другое утверждение."""
        client.force_login(make_user(pages=[PAGE_KEY]))

        body = client.get(URL).json()

        assert all(row["deferral_days"] is None for row in body["results"])
        assert body["coverage"]["with_deferral_count"] == 0

    def test_detail_answers_by_counterparty(self, client, make_user, owing):
        client.force_login(make_user(pages=[PAGE_KEY]))

        body = client.get(f"{URL}{owing.id}/").json()

        assert body["documents_count"] == 2
        assert body["debt_kopecks"] == 300_000
        assert body["oldest_age_days"] == 70

    def test_detail_of_a_stranger_is_404(self, client, make_user, owing):
        client.force_login(make_user(pages=[PAGE_KEY]))
        assert client.get(f"{URL}10000/").status_code == 404


class TestExport:
    def test_file_carries_two_sheets(self, client, make_user, owing):
        """Второй лист — сами документы: вопрос «за что именно» задают первым."""
        from io import BytesIO

        from openpyxl import load_workbook

        client.force_login(make_user(pages=[PAGE_KEY]))

        response = client.get(f"{URL}xlsx/")

        assert response.status_code == 200
        book = load_workbook(BytesIO(b"".join(response.streaming_content)))
        assert book.sheetnames == ["Контрагенты", "Документы"]
        # Шапка, двое должников, площадка и строка итога: площадка в итог
        # не входит, но из файла не исчезает — товар ушёл, деньги не пришли.
        assert book["Контрагенты"].max_row == 5
        # На втором листе — и товар по комиссии тоже: вопрос «а где остальные
        # отгрузки Каприоля» задаётся первым, и файл обязан отвечать.
        assert book["Документы"].max_row == 6
