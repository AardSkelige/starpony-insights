"""Выгрузка «Сроки оплаты»: два листа и колонка доли.

Дефект, ради которого написан этот файл, на экране не виден вовсе. Дебиторка
и площадки стоят там порознь, и каждая считает долю внутри себя; в файле они
попадают в одну колонку — и доли складываются в двести процентов, ровно там,
где заметить это невозможно.
"""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from api.deadlines.services import deadlines as service, excel

pytestmark = pytest.mark.django_db

# Колонка «Доля в долге» — четвёртая: контрагент, расчёты, долг, доля.
SHARE_COLUMN = 3


@pytest.fixture
def mixed(make_agent, make_document):
    buyer = make_agent("ООО «ПМТ»")
    other = make_agent("КРМОО «Каприоль»")
    ozon = make_agent("ООО «Интернет Решения»", tags=["маркетплейсы"])

    make_document(agent=buyer, total_kopecks=300_000)
    make_document(agent=other, total_kopecks=100_000)
    make_document(agent=ozon, total_kopecks=900_000)
    return buyer


def sheet_rows(filters: service.Filters):
    book = load_workbook(BytesIO(excel.build(filters).read()))
    sheet = book["Контрагенты"]
    return [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]


class TestShares:
    def test_marketplace_has_no_share_in_the_file(self, mixed):
        """Доля площадки в дебиторке не определена — площадка в неё не входит.

        Пустая ячейка честнее нуля и честнее числа: ноль читался бы как
        «ничего не должны», а доля внутри площадок — как доля в долге.
        """
        rows = sheet_rows(service.Filters())
        by_name = {row[0]: row for row in rows}

        assert by_name["ООО «Интернет Решения»"][SHARE_COLUMN] is None

    def test_shares_of_the_file_add_up_to_one(self, mixed):
        """Сложение колонки даёт ровно сто процентов, а не двести."""
        rows = sheet_rows(service.Filters())
        shares = [row[SHARE_COLUMN] for row in rows if row[SHARE_COLUMN] is not None]

        # Последняя строка — итог; он тоже несёт долю и в сложение не входит.
        assert sum(shares[:-1]) == pytest.approx(1.0)

    def test_marketplaces_stay_in_the_file(self, mixed):
        """Доли у площадки нет, но сама она из файла не исчезает:
        товар ушёл, деньги не пришли."""
        names = [row[0] for row in sheet_rows(service.Filters())]

        assert "ООО «Интернет Решения»" in names
