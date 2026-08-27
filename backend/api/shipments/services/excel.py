"""Выгрузка таблицы в XLSX.

Отдельным модулем от расчёта: сводка по товарам нужна и экрану, и файлу,
а знание про ширину колонок и формат ячейки — только файлу.

**Выгружается вся выборка, а не видимая страница.** Человек нажимает
«Экспорт», отобрав период и канал, и ждёт в файле то, что отобрал, —
а не пятьдесят строк, случайно попавших на первый экран.
"""

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from api.shipments.services import products

# Заголовок, ширина колонки и как достать значение из строки сводки.
# Порядок и подписи совпадают с экраном: файл, где колонки идут иначе,
# заставляет сверять их глазами.
COLUMNS: tuple[tuple[str, int, str], ...] = (
    ("Код", 12, "code"),
    ("Артикул", 14, "article"),
    ("Наименование", 52, "name"),
    ("Ед.", 6, "uom"),
    ("Продано", 12, "quantity"),
    ("в т.ч. даром", 13, "free_quantity"),
    ("Выручка, ₽", 15, "revenue"),
    ("Средняя цена продажи, ₽", 22, "avg_price"),
    ("Без учёта бесплатных, ₽", 22, "avg_price_paid"),
    ("Доля в выручке", 15, "share"),
    ("Отгрузок", 10, "documents_count"),
)

MONEY = "# ##0.00"
QUANTITY = "# ##0.###"
SHARE = "0.0%"

# Excel считает в числах с плавающей точкой, и точность Decimal ему не
# передать. Поэтому в файл уходят рубли числом: он для чтения и сводных
# таблиц, а сверка до копейки идёт по экрану и по самому учёту.
FORMATS = {
    "quantity": QUANTITY,
    "free_quantity": QUANTITY,
    "revenue": MONEY,
    "avg_price": MONEY,
    "avg_price_paid": MONEY,
    "share": SHARE,
}


def _cell_values(row: dict) -> dict:
    return {
        "code": row["code"],
        "article": row["article"],
        "name": row["name"],
        "uom": row["uom"],
        "quantity": float(row["quantity"]),
        "free_quantity": float(row["free_quantity"]),
        "revenue": float(Decimal(row["revenue_kopecks"]) / 100),
        "avg_price": _rubles(row["avg_price_kopecks"]),
        "avg_price_paid": _rubles(row["avg_price_paid_kopecks"]),
        "share": float(row["revenue_share"]) if row["revenue_share"] is not None else None,
        "documents_count": row["documents_count"],
    }


def _rubles(kopecks: Decimal | None) -> float | None:
    return float(kopecks / 100) if kopecks is not None else None


def build(filters: products.Filters) -> BytesIO:
    """Собрать книгу по всей выборке фильтров."""
    book = Workbook()
    sheet = book.active
    sheet.title = "Товары в отгрузках"

    for index, (title, width, _) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width

    # Страницы обходятся пачками: держать в памяти всю выборку незачем,
    # а у приёмок строк будет заметно больше, чем шестьдесят шесть.
    totals = products.summary(filters)

    line = 2
    for row in _every_row(filters, totals["revenue_kopecks"]):
        values = _cell_values(row)
        for index, (_, __, key) in enumerate(COLUMNS, start=1):
            cell = sheet.cell(row=line, column=index, value=values[key])
            if key in FORMATS:
                cell.number_format = FORMATS[key]
        line += 1

    _write_totals(sheet, line, totals)

    # Шапка закрепляется: без этого на сороковой строке уже не видно,
    # что за колонка, и файл читают, прокручивая туда-обратно.
    sheet.freeze_panes = "A2"

    stream = BytesIO()
    book.save(stream)
    stream.seek(0)
    return stream


def _every_row(filters: products.Filters, total_revenue: int):
    """Все строки выборки, пачками.

    Идёт по срезам запроса напрямую, а не через `products.page`: та на каждый
    вызов пересчитывает итоги и общее число строк, и на выгрузке из тысячи
    позиций это пять лишних проходов по всем документам.
    """
    from dataclasses import replace

    selection = products.grouped(filters)
    start = 0
    while True:
        chunk = list(selection[start : start + products.MAX_PAGE_SIZE])
        if not chunk:
            return
        for item in chunk:
            yield products.row_of(item, total_revenue)
        if len(chunk) < products.MAX_PAGE_SIZE:
            return
        start += products.MAX_PAGE_SIZE


def _write_totals(sheet, line: int, totals: dict) -> None:
    values = {
        "name": f"Итого · {totals['products_count']} наименований",
        "quantity": float(totals["quantity"]),
        "free_quantity": float(totals["free_quantity"]),
        "revenue": float(Decimal(totals["revenue_kopecks"]) / 100),
        "documents_count": totals["documents_count"],
    }
    for index, (_, __, key) in enumerate(COLUMNS, start=1):
        if key not in values:
            continue
        cell = sheet.cell(row=line, column=index, value=values[key])
        cell.font = Font(bold=True)
        if key in FORMATS:
            cell.number_format = FORMATS[key]
