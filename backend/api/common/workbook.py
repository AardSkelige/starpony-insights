"""Каркас книги XLSX: листы, шапка, ширины, формат ячейки, строка итога.

Общее для всех выгрузок. Знание про то, что за числа в колонках, остаётся
у страницы; здесь только про то, как книга устроена.

**Выгружается вся выборка, а не видимая страница.** Человек нажимает
«Экспорт», отобрав период и канал, и ждёт в файле то, что отобрал, —
а не пятьдесят строк, случайно попавших на первый экран.
"""

from dataclasses import dataclass, field
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

MONEY = "# ##0.00"
QUANTITY = "# ##0.###"
UNIT_PRICE = "# ##0.0000"
SHARE = "0.0%"

# Заголовок, ширина колонки и ключ значения в строке.
Column = tuple[str, int, str]


@dataclass
class Sheet:
    """Один лист книги."""

    title: str
    columns: tuple[Column, ...]
    rows: Iterable[dict]
    formats: dict[str, str] = field(default_factory=dict)
    totals: dict | None = None


def build(*sheets: Sheet) -> BytesIO:
    """Собрать книгу из готовых строк-словарей."""
    book = Workbook()
    # Первый лист создаёт сам openpyxl — им и пользуемся, иначе в книге
    # остаётся пустой «Sheet» перед нашими данными.
    target = book.active

    for index, sheet in enumerate(sheets):
        if index > 0:
            target = book.create_sheet()
        _fill(target, sheet)

    stream = BytesIO()
    book.save(stream)
    stream.seek(0)
    return stream


def _fill(target, sheet: Sheet) -> None:
    # Excel режет имя листа на 31 знаке и падает на символах `:\/?*[]`.
    target.title = sheet.title[:31]

    for index, (header, width, _) in enumerate(sheet.columns, start=1):
        cell = target.cell(row=1, column=index, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        target.column_dimensions[get_column_letter(index)].width = width

    line = 2
    for row in sheet.rows:
        _write(target, line, sheet, row)
        line += 1

    if sheet.totals is not None:
        _write(target, line, sheet, sheet.totals, bold=True)

    # Шапка закрепляется: без этого на сороковой строке уже не видно,
    # что за колонка, и файл читают, прокручивая туда-обратно.
    target.freeze_panes = "A2"


def _write(target, line: int, sheet: Sheet, values: dict, bold: bool = False) -> None:
    for index, (_, __, key) in enumerate(sheet.columns, start=1):
        if key not in values:
            continue
        cell = target.cell(row=line, column=index, value=values[key])
        if bold:
            cell.font = Font(bold=True)
        if key in sheet.formats:
            cell.number_format = sheet.formats[key]
